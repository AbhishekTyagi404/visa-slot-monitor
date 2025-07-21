import time
import os
import csv
import io
import smtplib
from email.mime.text import MIMEText
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.service import Service
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from datetime import datetime
import difflib

# ==== CONFIG ====
URL = "https://checkvisaslots.com/latest-us-visa-availability.html#F1Regular"
EXCEL_PATH = "visa_slot_log.xlsx"
TEXT_LOG = "text.txt"

SENDER_EMAIL = "sender-email@gmail.com"
SENDER_PASSWORD = "pjkronwbenrslfdh"
RECEIVER_EMAILS = [
    "reciever-email1@gmail.com",
    "reciever-email2@gmail.com"
]

COLUMNS_TO_KEEP = ['visa location', 'visa type', 'last seen at']

# ==== EMAIL ALERT FUNCTION ====
def send_email_alert(changed_rows):
    first_row = changed_rows[0]
    subject = f"Visa Slot Changed: {first_row[0]} at {first_row[2]}"

    body_lines = ["Changes detected in visa slots:\n"]
    for row in changed_rows:
        body_lines.append(", ".join(row))
    body = "\n".join(body_lines)

    msg = MIMEText(body)
    msg['From'] = SENDER_EMAIL
    msg['To'] = ", ".join(RECEIVER_EMAILS)
    msg['Subject'] = subject

    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.sendmail(SENDER_EMAIL, RECEIVER_EMAILS, msg.as_string())
        server.quit()
        print("📧 Email alert sent.")
    except Exception as e:
        print(f"❌ Failed to send email: {e}")


# ==== EXCEL FUNCTIONS ====
def initialize_excel():
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "VisaSlots"
        ws.append(["Timestamp"] + [col.title() for col in COLUMNS_TO_KEEP])
        wb.save(EXCEL_PATH)

def append_to_excel(rows):
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    for row in rows:
        ws.append([timestamp] + row)

    wb.save(EXCEL_PATH)


# ==== TEXT LOG FUNCTIONS ====
def read_old_text():
    if os.path.exists(TEXT_LOG):
        with open(TEXT_LOG, 'r', encoding='utf-8') as f:
            return f.read()
    return ""

def save_text(text):
    with open(TEXT_LOG, 'w', encoding='utf-8') as f:
        f.write(text)

def get_rows_from_csv(csv_text):
    reader = csv.reader(io.StringIO(csv_text))
    next(reader, None)  # Skip header
    return [row for row in reader]


# ==== TABLE PARSING ====
def get_f1_regular_table_html(driver):
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    header = soup.find(lambda tag: tag.name in ["h2", "h3", "h4", "strong"] and "Last F-1 (Regular) Availability" in tag.text)
    if not header:
        print("❌ Header 'Last F-1 (Regular) Availability' not found.")
        return None

    next_el = header.find_next_sibling()
    while next_el and next_el.name != 'table':
        next_el = next_el.find_next_sibling()
    if not next_el:
        print("❌ Table after header not found.")
        return None

    return next_el

def extract_relevant_columns(table_tag):
    rows = table_tag.find_all('tr')
    if not rows:
        return [], ""

    headers = [th.get_text(strip=True).lower() for th in rows[0].find_all(['th', 'td'])]
    indexes = []
    for col in COLUMNS_TO_KEEP:
        try:
            indexes.append(headers.index(col))
        except ValueError:
            print(f"❌ Column '{col}' not found.")
            return [], ""

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([COLUMNS_TO_KEEP[i].title() for i in range(len(COLUMNS_TO_KEEP))])

    extracted_rows = []
    for row in rows[1:]:
        cells = row.find_all(['td', 'th'])
        if len(cells) < max(indexes) + 1:
            continue
        selected_cells = [cells[i].get_text(strip=True) for i in indexes]
        extracted_rows.append(selected_cells)
        writer.writerow(selected_cells)

    return extracted_rows, output.getvalue()


# ==== MAIN LOOP ====
def main():
    options = Options()
    # options.add_argument('--headless')  # Uncomment to run in background

    service = Service()
    driver = webdriver.Firefox(service=service, options=options)

    try:
        initialize_excel()
        driver.get(URL)
        old_text = read_old_text()
        old_rows = get_rows_from_csv(old_text) if old_text else []

        while True:
            time.sleep(3)
            driver.refresh()
            time.sleep(2)

            table_tag = get_f1_regular_table_html(driver)
            if table_tag is None:
                continue

            new_rows, new_csv = extract_relevant_columns(table_tag)
            if not new_csv:
                continue

            if new_csv != old_text:
                print("⚠️ Change detected!")

                changed_rows = [row for row in new_rows if row not in old_rows]
                if changed_rows:
                    print("\n🆕 New rows:")
                    for row in changed_rows:
                        print(" → ", row)

                    send_email_alert(changed_rows)
                    append_to_excel(changed_rows)
                else:
                    print("⚠️ Change detected, but no unique new rows found.")

                save_text(new_csv)
                old_text = new_csv
                old_rows = new_rows
            else:
                print("✅ No change detected.")

    except KeyboardInterrupt:
        print("\n⛔ Script stopped by user.")
    finally:
        driver.quit()

if __name__ == "__main__":
    main()
